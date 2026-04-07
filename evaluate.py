import csv
import asyncio
import json
import logging
import os

from agent import orchestrate_request, MODEL_PROVIDER
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

logging.getLogger("google_genai.types").setLevel(logging.ERROR)


def normalize_label(label):
    normalized = (label or "").strip().lower()
    if normalized in {"shopping", "queries"}:
        return "service"
    if normalized in {"guardrail", "unsafe"}:
        return "out_of_scope"
    if normalized in {"greeting", "device_control", "service", "automations", "out_of_scope"}:
        return normalized
    return "out_of_scope"


def save_pretty_excel(matrix, labels):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = "MISCLASSIFICATION MATRIX"
    ws["A1"].fill = header_fill
    ws["A1"].font = bold
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Model: {MODEL_PROVIDER} | Rows = Actual | Columns = Predicted | Green = Correct"
    ws["A2"].alignment = center

    ws.merge_cells("A4:F4")
    ws["A4"] = "Orchestrator Evaluation"
    ws["A4"].fill = subheader_fill
    ws["A4"].alignment = center

    ws.cell(row=5, column=1, value="Actual ↓ Predicted →").alignment = center

    for j, label in enumerate(labels, start=2):
        ws.cell(row=5, column=j, value=label).alignment = center

    ws.cell(row=5, column=len(labels) + 2, value="Total")

    for i, actual in enumerate(labels, start=6):
        ws.cell(row=i, column=1, value=actual)

        row_total = 0
        for j, pred in enumerate(labels, start=2):
            value = matrix[actual].get(pred, 0)
            row_total += value

            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = center

            if actual == pred:
                cell.fill = green
            elif value > 0:
                cell.fill = red

        ws.cell(row=i, column=len(labels) + 2, value=row_total)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    wb.save("orchestration_metrics.xlsx")
    print("✅ orchestration_metrics.xlsx created")


async def main():
    if not os.path.exists("dataset.csv"):
        print("❌ Error: dataset.csv not found.")
        return

    with open("dataset.csv", "r", encoding="utf-8") as f:
        dataset = list(csv.DictReader(f))

    if os.path.exists("progress.json") or os.path.exists("evaluation_results.csv"):
        choice = input("Old results found. Reset evaluation? (y/n): ").lower()
        if choice == "y":
            if os.path.exists("progress.json"):
                os.remove("progress.json")
            if os.path.exists("evaluation_results.csv"):
                os.remove("evaluation_results.csv")
            print("✨ Progress cleared. Starting fresh...")
        else:
            print("⏩ Resuming previous evaluation...")

    start_index = 0
    if os.path.exists("progress.json"):
        with open("progress.json", "r", encoding="utf-8") as f:
            progress = json.load(f)
            start_index = progress.get("last_index", 0)

    actual = []
    predicted = []

    if os.path.exists("evaluation_results.csv"):
        with open("evaluation_results.csv", "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                actual.append(normalize_label(row["Actual Label"]))
                predicted.append(normalize_label(row["Predicted Label"]))

    print(f"Testing model: {MODEL_PROVIDER}")
    print(f"Total entries: {len(dataset)} | Starting at: {start_index}\n")

    try:
        for i in range(start_index, len(dataset)):
            item = dataset[i]
            text = item["input_text"]
            label = normalize_label(item["actual_label"])

            try:
                result = await orchestrate_request(text)
                pred = normalize_label(result.get("intent", "out_of_scope"))
            except Exception as e:
                print(f"❌ Error at index {i}: {e}")
                pred = "error"

            actual.append(label)
            predicted.append(pred)

            print(f"[{i + 1}/{len(dataset)}] Query: {text}")
            print(f"Result -> Actual: {label}, Predicted: {pred}\n")

            file_exists = os.path.exists("evaluation_results.csv")
            with open("evaluation_results.csv", "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(["Input Query", "Actual Label", "Predicted Label", "Correct"])
                writer.writerow([text, label, pred, label == pred])

            with open("progress.json", "w", encoding="utf-8") as f:
                json.dump({"last_index": i + 1}, f)

            await asyncio.sleep(0.2)

    except Exception as e:
        print("⚠️ Evaluation stopped due to error:", e)

    finally:
        if not actual:
            print("No results to report.")
            return

        print("\n📊 Generating reports...")

        valid_labels = ["device_control", "greeting", "service", "automations", "out_of_scope", "error"]
        unique_labels = sorted(list(set(actual + predicted + valid_labels)))

        matrix = {a: {p: 0 for p in unique_labels} for a in unique_labels}
        for a, p in zip(actual, predicted):
            matrix[a][p] += 1

        print("\n================== Confusion Matrix ==================")
        header = f"{'Actual \\ Pred':<20} | " + " | ".join([f"{l[:10]:<10}" for l in unique_labels])
        print(header)
        print("-" * len(header))

        for a in unique_labels:
            row = f"{a:<20} | "
            for p in unique_labels:
                row += f"{matrix[a][p]:<10} | "
            print(row)

        correct = sum(1 for a, p in zip(actual, predicted) if a == p)
        total = len(actual)
        print(f"\nAccuracy: {correct}/{total} ({(correct / total) * 100:.2f}%)")

        with open("confusion_matrix.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Actual \\ Predicted"] + unique_labels + ["Total"])
            for a in unique_labels:
                row = [a]
                total_row = 0
                for p in unique_labels:
                    val = matrix[a][p]
                    row.append(val)
                    total_row += val
                row.append(total_row)
                writer.writerow(row)

        save_pretty_excel(matrix, unique_labels)
        print("✅ Reports generated successfully.")


if __name__ == "__main__":
    asyncio.run(main())